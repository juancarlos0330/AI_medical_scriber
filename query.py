import tkinter as tk
import pymongo  
from fuzzywuzzy import fuzz, process
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import requests
import os
from transformers import pipeline
import fitz  # Import PyMuPDF
import json

client = pymongo.MongoClient("mongodb+srv://arbab998:wDl85YMTveO5ayrQ@cluster0.tjglijc.mongodb.net/")
db=client["medical_data"]
collection = db["timeline"]
threshold = 95

selected_patient = ''
selected_doctor = ''

API_URL = "https://api-inference.huggingface.co/models/deepset/roberta-base-squad2"
headers = {"Authorization": "Bearer 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"}
text_path = 'json/text'
context = ""
question = ""
count = 0

img_pdf_path = 'query_resource/pdf_images'
img_text_path = 'query_resource/pdf_jsons'
arzure_api_url = "https://portal.vision.cognitive.azure.com/api/demo/analyze?features=read"
src_path = 'query_resource/pdf_documents'
img_folder_path = ""
json_folder_path = ""
selected_pdf = ""
text_path = 'query_resource/pdf_texts'

def ai_proc():
  qa_pipeline = pipeline("question-answering", model="deepset/roberta-base-squad2")

  # Define the context from your text
  context_txt = ""

  with open(os.path.join(text_path, selected_pdf[:-4]+".txt"), 'r', encoding='utf-8-sig') as ff:
      context_txt = ff.read() 

  # List of questions to get answers for
  questions = [
      "Who is author?",
      "Who was the document electronically signed?",
      "When was the document electronically signed?",
      "Where was the patient treated?",
      "what are the services?",
      "What is the patient's date of birth?"
  ]

  # Loop through each question, get the answer, and print it
  answers=[]
  for question in questions:
    result = qa_pipeline(question=question, context=context_txt)
    print(f"Question: {question}\nAnswer: {result['answer']}\n")
    answers.append(result['answer'])
  
  return {"questions": questions, "result": answers}


def proc_data(data, filename):
  for page in data:
    blocks_data = page["readResult"]["blocks"]
    for block in blocks_data:
      lines_data = block["lines"]
      for line in lines_data:
        text = line["text"]
        with open(os.path.join(text_path, filename)+'.txt', 'a', encoding='utf-8') as ff:
          ff.write(text + '\n')

def ocr(src, dst):
  try:
    file_name = os.path.basename(src)

    with open(src, 'rb') as f:
      files=[('file',(file_name,f,'image/png'))]
      response = requests.request("POST", arzure_api_url, data={}, files=files)
      # print(json.loads(response.text)["modelVersion"])
      temp_file = os.path.join(dst+'/', file_name[:-4] + '.json')
      with open(temp_file, 'w', encoding='utf-8-sig') as ff:
        json.dump(json.loads(response.text), ff, indent=4)

    return
  except Exception as err:
      print('Error:', err)
      return

def extract(pdf_file_path):
  print("processing ", pdf_file_path)

  pdf_document = fitz.open(pdf_file_path)

  # Define your desired output width and height
  desired_width = 1654  # example width in pixels
  desired_height = 2339  # example height in pixels

  for page_num in range(len(pdf_document)):
    page = pdf_document[page_num]
    
    # Original dimensions of the PDF page
    orig_width, orig_height = page.rect.width, page.rect.height

    # Calculate scale factors to achieve desired dimensions
    scale_x = desired_width / orig_width
    scale_y = desired_height / orig_height

    # Create a matrix for the transformation
    mat = fitz.Matrix(scale_x, scale_y)

    # Generate the pixmap with the transformation
    pix = page.get_pixmap(matrix=mat)
    
    output_image_path = os.path.join(img_folder_path+'/', 'page{:03d}.png'.format(page_num))
    pix.save(output_image_path)

  pdf_document.close()

  temp_img_files = os.listdir(img_folder_path)
  for temp_img_file in temp_img_files:
    temp_img_file_path = os.path.join(img_folder_path+'/', temp_img_file)
    ocr(temp_img_file_path, json_folder_path)
    
def ask_question(context, question):
    payload = {"inputs": {"question": question, "context": context}}
    response = requests.post(API_URL, headers=headers, json=payload)
    return response.json()

# Function to be called when query about the doctor
def on_query_doctor():
  global selected_patient
  selected_indices = listbox.curselection()
  if selected_indices:  # If there is at least one selected item
    selected_index = selected_indices[0]  # Get the first selected item's index
    selected_patient = listbox.get(selected_index)  # Get the item's text
    label_sel_patient.config(text=selected_patient)

    query = {'patient_name': selected_patient}
    distinct_doctors = collection.distinct('doctors', query)
    listbox_doctors.delete(0, tk.END)
    similar_groups = []
    checked = []
    
    for doctor in distinct_doctors:
      if doctor not in checked:
        # Find matches in the list itself, with a similarity above the threshold
        matches = process.extract(doctor, distinct_doctors, limit=None, scorer=fuzz.token_sort_ratio)
        # Filter matches by the threshold, excluding the current patient (100% match)
        similar = [match for match, score in matches if score >= threshold and match != doctor]
        if similar:
            similar.append(doctor)  # Add the current patient to the group
            similar_groups.append(similar)  # Add this group to the list of groups
            checked.extend(similar)  # Mark these patients as checked
        else:
            checked.append(doctor)  # Mark this patient as checked if no similar found
    
    # Flatten array2 into a single list of elements to remove
    elements_to_remove = [element for sublist in similar_groups for element in sublist]
    # Remove elements from array1 that are present in elements_to_remove
    filtered_array = [element for element in distinct_doctors if element not in elements_to_remove]

    for item in similar_groups:
      listbox_doctors.insert(tk.END, item[0])
    for item in filtered_array:
      listbox_doctors.insert(tk.END, item)

# Function to be called when query about the date
def on_query_date():
  global selected_doctor
  selected_indices_doctor = listbox_doctors.curselection()

  if selected_patient and selected_indices_doctor:  # If there is at least one selected item
    selected_index_doctor = selected_indices_doctor[0]
    selected_doctor = listbox_doctors.get(selected_index_doctor)
    label_sel_doctor.config(text=selected_doctor)

    query = {'patient_name': selected_patient, 'doctors': selected_doctor}
    distinct_dates = collection.distinct('date', query)
    listbox_date.delete(0, tk.END)
    
    for date in distinct_dates:
      date_string = date.strftime('%m/%d/%Y')
      listbox_date.insert(tk.END, date_string)

# Function to be called when query about the activity
def on_query_activity():
  selected_indices_date = listbox_date.curselection()
  if selected_patient and selected_doctor and selected_indices_date:  # If there is at least one selected item
    selected_index_date = selected_indices_date[0]
    selected_date = listbox_date.get(selected_index_date)
    label_sel_date.config(text=selected_date)

    query = {'patient_name': selected_patient, 'doctors': selected_doctor, 'date': datetime.strptime(selected_date, '%m/%d/%Y')}
    distinct_activites = collection.distinct('activity', query)
    listbox_activity.delete(0, tk.END)
    
    for activity in distinct_activites:
      listbox_activity.insert(tk.END, activity)

#write excel file
def write_excel():
  wb = openpyxl.Workbook()
  distinct_patients = collection.distinct('patient_name')
  custom_font = Font(size=14, bold=True)
  for patient in distinct_patients:
    ws = wb.create_sheet(patient)
    ws['A1'].font = custom_font
    ws['B1'].font = custom_font
    ws['C1'].font = custom_font

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 60

    ws['A1'] = 'Date'
    ws['B1'] = 'Treatment Timeline'
    ws['C1'] = 'Doctors Name'

    results = collection.find({'patient_name': patient}).sort([
      ('date', pymongo.ASCENDING)
    ])
    
    for res in results:
      doctors = ''
      for w_doctor in res['doctors']:
        doctors = doctors + w_doctor + ' / '
      
      date_string = res['date'].strftime('%m/%d/%Y')
      doc = [date_string, res['activity'], doctors]
      ws.append(doc)
    
  wb.save("Timeline_uni.xlsx")
  print("made excel file")

def on_question():
  global context
  global img_folder_path
  global json_folder_path
  global selected_pdf

  selected_indices_pdf = listbox_message.curselection()
  selected_index_pdf = selected_indices_pdf[0]
  selected_pdf = listbox_message.get(selected_index_pdf)

  if not os.path.exists(os.path.join(img_pdf_path+'/',selected_pdf[:-4])):
    print("File does not exist")
    img_folder_path = os.path.join(img_pdf_path+'/', selected_pdf[:-4])
    json_folder_path = os.path.join(img_text_path+'/', selected_pdf[:-4])
    os.makedirs(img_folder_path, exist_ok=True)
    os.makedirs(json_folder_path, exist_ok=True)
    pdf_file_path = os.path.join(src_path, selected_pdf)
    extract(pdf_file_path)
  
  if not os.path.exists(os.path.join(text_path+'/',selected_pdf[:-4]+'.txt')):
    print("File does not exist")
    temp_json_files = os.listdir(os.path.join(img_text_path+'/', selected_pdf[:-4]))
    data = []
    for temp_json_file in temp_json_files:
      try:
        with open(os.path.join(os.path.join(img_text_path+'/', selected_pdf[:-4]), temp_json_file), 'r', encoding='utf-8-sig') as json_file:
          temp = json_file.read()
          if temp:  # Ensure 'temp' is not empty
            data.append(json.loads(temp))
          else:
            print("Warning: The file is empty.")
      except FileNotFoundError:
        print("Error: The file was not found.")
      except json.JSONDecodeError as e:
        print(f"Error decoding JSON: {e}")
      
    proc_data(data, selected_pdf[:-4])  
  
  answers = ai_proc()
  questions = answers["questions"]
  results = answers["result"]

  input_answer.delete('1.0', 'end')

  for i in range(0, len(questions)):
    input_answer.insert(tk.END, questions[i] + '\n') 
    input_answer.insert(tk.END, results[i] + '\n') 
    input_answer.insert(tk.END, '\n') 

# Create the main window
root = tk.Tk()
root.title("Medical scribe expert")
root.geometry("850x410")  # Set the size of the window

# list patients
label = tk.Label(root, text="Patients List")
label.pack(pady=10)  # Add some padding around the label
label.place(x=25, y=25)

distinct_patients = collection.distinct('patient_name')

listbox = tk.Listbox(root)
listbox.pack(pady=5)
listbox.place(x=15, y=60)

for item in distinct_patients:
    listbox.insert(tk.END, item)

# query doctor
label_q1 = tk.Label(root, text="Which doctors saw the patient?")
label_q1.pack(pady=10)  # Add some padding around the label
label_q1.place(x=170, y=5)

label_d1 = tk.Label(root, text="Select the patient.")
label_d1.pack(pady=10)  # Add some padding around the label
label_d1.place(x=170, y=35)

button_doc = tk.Button(root, text=">>", command=on_query_doctor)
button_doc.pack(pady=5)  # Add some padding around the button
button_doc.place(x=300, y=35, height=20)

listbox_doctors = tk.Listbox(root)
listbox_doctors.pack(pady=5)
listbox_doctors.place(x=175, y=60, width=200, height=170)

# next query  (date)
label_q2 = tk.Label(root, text="when the doctor saw the patient?")
label_q2.pack(pady=10)  # Add some padding around the label
label_q2.place(x=420, y=5)

label_d2 = tk.Label(root, text="Select the doctor.")
label_d2.pack(pady=10)  # Add some padding around the label
label_d2.place(x=420, y=35)

button_date = tk.Button(root, text=">>", command=on_query_date)
button_date.pack(pady=5)  # Add some padding around the button
button_date.place(x=550, y=35, height=20)

listbox_date = tk.Listbox(root)
listbox_date.pack(pady=5)
listbox_date.place(x=425, y=60, width=150, height=170)

# next query  (activity)
label_q3 = tk.Label(root, text="what did the doctor?")
label_q3.pack(pady=10)  # Add some padding around the label
label_q3.place(x=620, y=5)

label_d3 = tk.Label(root, text="Select the date.")
label_d3.pack(pady=10)  # Add some padding around the label
label_d3.place(x=620, y=35)

button_act = tk.Button(root, text=">>", command=on_query_activity)
button_act.pack(pady=5)  # Add some padding around the button
button_act.place(x=790, y=35, height=20)

listbox_activity = tk.Listbox(root)
listbox_activity.pack(pady=5)
listbox_activity.place(x=625, y=60, width=200, height=170)

#display selected items
label_patient = tk.Label(root, text="Selected Patient :")
label_patient.pack(pady=10)  # Add some padding around the label
label_patient.place(x=30, y=250)

label_sel_patient = tk.Label(root, text="")
label_sel_patient.pack(pady=10)  # Add some padding around the label
label_sel_patient.place(x=130, y=250)

label_doctor = tk.Label(root, text="Selected Doctor :")
label_doctor.pack(pady=10)  # Add some padding around the label
label_doctor.place(x=280, y=250)

label_sel_doctor = tk.Label(root, text="")
label_sel_doctor.pack(pady=10)  # Add some padding around the label
label_sel_doctor.place(x=380, y=250)

label_date = tk.Label(root, text="Selected Date :")
label_date.pack(pady=10)  # Add some padding around the label
label_date.place(x=600, y=250)

label_sel_date = tk.Label(root, text="")
label_sel_date.pack(pady=10)  # Add some padding around the label
label_sel_date.place(x=700, y=250)

#gpt query
label_message = tk.Label(root, text="pdf documents of the patience")
label_message.pack(pady=10)  # Add some padding around the label
label_message.place(x=30, y=290)

listbox_message = tk.Listbox(root)
listbox_message.pack(pady=5)
listbox_message.place(x=30, y=315, width=360, height=85)

pdf_files = os.listdir(src_path)
for pdf_file in pdf_files:
  pdf_file_path = os.path.join(src_path, pdf_file)
  listbox_message.insert(tk.END, pdf_file)

button_message = tk.Button(root, text=">>", command=on_question)
button_message.pack(pady=5)  # Add some padding around the button
button_message.place(x=420, y=350, height=20)

label_answer = tk.Label(root, text="answer")
label_answer.pack(pady=10)  # Add some padding around the label
label_answer.place(x=480, y=290)

input_answer = tk.Text(root)
input_answer.place(x=480, y=315, width=300, height=85)
#excel
# button_excel = tk.Button(root, text="Excel", command=write_excel)
# button_excel.pack(pady=5)  # Add some padding around the button
# button_excel.place(x=800, y=250, height=20)
# Start the GUI event loop
root.mainloop()
